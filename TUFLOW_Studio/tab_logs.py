import os
from qgis.PyQt.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton,
    QLineEdit, QFileDialog, QTableWidget, QTableWidgetItem,
    QProgressBar, QHeaderView, QAbstractItemView, QMessageBox,
)
from qgis.PyQt.QtCore import Qt, QThread, pyqtSignal
from qgis.PyQt.QtGui import QColor
from .tlf_parser import scan_log_dir

HEADERS = ['File Name', 'Scenarios', 'Events', 'Simulation Date',
           'Simulation Time', 'Duration', 'End Time']


class _ScanWorker(QThread):
    progress = pyqtSignal(int)
    finished = pyqtSignal(list)
    error = pyqtSignal(str)

    def __init__(self, log_dir):
        super().__init__()
        self.log_dir = log_dir

    def run(self):
        try:
            results = scan_log_dir(self.log_dir)
            self.finished.emit(results)
        except Exception as e:
            self.error.emit(str(e))


class TabLogs(QWidget):
    busy_changed = pyqtSignal(bool)

    def __init__(self, parent=None):
        super().__init__(parent)
        self._results = []
        self._worker = None
        self._build_ui()

    def _build_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(8, 8, 8, 8)

        # Log dir row
        dir_row = QHBoxLayout()
        dir_row.addWidget(QLabel('Log folder:'))
        self._dir_edit = QLineEdit()
        self._dir_edit.setPlaceholderText('runs\\log\\  (auto-filled from Root tab)')
        dir_row.addWidget(self._dir_edit)
        btn_browse = QPushButton('Browse…')
        btn_browse.clicked.connect(self._browse_log_dir)
        dir_row.addWidget(btn_browse)
        layout.addLayout(dir_row)

        # Buttons row
        btn_row = QHBoxLayout()
        btn_scan = QPushButton('Scan')
        btn_scan.clicked.connect(self._start_scan)
        self._btn_export = QPushButton('Export to Excel')
        self._btn_export.clicked.connect(self._export_excel)
        self._btn_export.setEnabled(False)
        btn_row.addWidget(btn_scan)
        btn_row.addWidget(self._btn_export)
        btn_row.addStretch()
        layout.addLayout(btn_row)

        # Progress bar
        self._progress = QProgressBar()
        self._progress.setVisible(False)
        self._progress.setRange(0, 0)  # indeterminate
        layout.addWidget(self._progress)

        # Status label
        self._status = QLabel('')
        layout.addWidget(self._status)

        # Table
        self._table = QTableWidget(0, len(HEADERS))
        self._table.setHorizontalHeaderLabels(HEADERS)
        self._table.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
        for col in range(1, len(HEADERS)):
            self._table.horizontalHeader().setSectionResizeMode(col, QHeaderView.ResizeToContents)
        self._table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self._table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self._table.setAlternatingRowColors(True)
        layout.addWidget(self._table)

    # ------------------------------------------------------------------
    # Actions
    # ------------------------------------------------------------------
    def _browse_log_dir(self):
        path = QFileDialog.getExistingDirectory(self, 'Select log folder')
        if path:
            self._dir_edit.setText(path)

    def set_log_dir(self, path):
        """Called by dock_widget when root paths change."""
        if path and os.path.isdir(path):
            self._dir_edit.setText(path)

    def _start_scan(self):
        log_dir = self._dir_edit.text().strip()
        if not log_dir or not os.path.isdir(log_dir):
            QMessageBox.warning(self, 'TUFLOW Studio', 'Please set a valid log folder first.')
            return
        self._progress.setVisible(True)
        self._status.setText('Scanning TLF files (Be patient please)…')
        self.busy_changed.emit(True)
        self._btn_export.setEnabled(False)
        self._table.setRowCount(0)
        self._worker = _ScanWorker(log_dir)
        self._worker.finished.connect(self._on_scan_done)
        self._worker.error.connect(self._on_scan_error)
        self._worker.start()

    def _on_scan_done(self, results):
        self._results = results
        self._progress.setVisible(False)
        self.busy_changed.emit(False)
        self._status.setText(f'Found {len(results)} run log(s).')
        self._populate_table(results)
        self._btn_export.setEnabled(bool(results))

    def _on_scan_error(self, msg):
        self._progress.setVisible(False)
        self._status.setText(f'Error: {msg}')

    def _populate_table(self, results):
        self._table.setRowCount(len(results))
        for r, row in enumerate(results):
            values = [row['file'], row['scenarios'], row['events'],
                      row['sim_date'], row['sim_time'], row['duration'], row['end_time']]
            for c, val in enumerate(values):
                item = QTableWidgetItem(val)
                item.setTextAlignment(Qt.AlignLeft | Qt.AlignVCenter)
                self._table.setItem(r, c, item)

    # ------------------------------------------------------------------
    # Excel export
    # ------------------------------------------------------------------
    def _export_excel(self):
        if not self._results:
            return
        path, _ = QFileDialog.getSaveFileName(
            self, 'Save Excel file', 'TUFLOW_Run_Log.xlsx',
            'Excel files (*.xlsx)')
        if not path:
            return
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment

            wb = Workbook()
            ws = wb.active
            ws.title = 'Run Log'

            header_font = Font(bold=True)
            header_fill = PatternFill(fill_type='solid', fgColor='D9D9D9')
            header_align = Alignment(horizontal='left')

            for c, h in enumerate(HEADERS, start=1):
                cell = ws.cell(row=1, column=c, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align

            keys = ['file', 'scenarios', 'events', 'sim_date', 'sim_time', 'duration', 'end_time']
            for r, row in enumerate(self._results, start=2):
                for c, key in enumerate(keys, start=1):
                    cell = ws.cell(row=r, column=c, value=row[key])
                    cell.alignment = Alignment(horizontal='left')
                    cell.number_format = '@'

            # Auto column width
            for col in ws.columns:
                max_len = max((len(str(cell.value or '')) for cell in col), default=0)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

            wb.save(path)
            QMessageBox.information(self, 'Export complete', f'Saved to:\n{path}')
        except ImportError:
            QMessageBox.critical(self, 'Error',
                'openpyxl is not available in this QGIS Python environment.\n'
                'Install it via: pip install openpyxl')
        except Exception as e:
            QMessageBox.critical(self, 'Export failed', str(e))
